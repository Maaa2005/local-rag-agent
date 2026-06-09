"""parser._parse_excel: 結合セル解除 → 同値補完 → Markdown 出力。"""
import tempfile
from pathlib import Path

import openpyxl
import pytest

from app.services.parser import _parse_excel


@pytest.fixture()
def xlsx_with_merged_cells():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Q1売上"
    ws["A1"] = "カテゴリ"
    ws["B1"] = "金額"
    ws["A2"] = "売上"
    ws["B2"] = 100
    ws["A3"] = None
    ws["B3"] = 200
    ws.merge_cells("A2:A3")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)
    wb.save(path)
    yield path
    path.unlink(missing_ok=True)


def test_parse_excel_includes_sheet_name(xlsx_with_merged_cells):
    md = _parse_excel(xlsx_with_merged_cells)
    assert "## シート: Q1売上" in md


def test_parse_excel_unmerges_and_fills(xlsx_with_merged_cells):
    """結合セル A2:A3 が解除されて両行とも「売上」になっていること。"""
    md = _parse_excel(xlsx_with_merged_cells)
    assert md.count("売上") >= 2
    assert "100" in md and "200" in md


def test_parse_excel_outputs_markdown_table(xlsx_with_merged_cells):
    md = _parse_excel(xlsx_with_merged_cells)
    assert "|" in md
