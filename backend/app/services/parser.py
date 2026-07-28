"""
ファイルパース: PDF/Word → Docling、Excel → openpyxl+pandas、テキスト → 直接読み込み。
Docling の重い AI モデルは CPU にオフロード（VRAM 保護）。
"""
import asyncio
from pathlib import Path


def _parse_excel(path: Path) -> str:
    import openpyxl
    import pandas as pd

    wb = openpyxl.load_workbook(path, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 結合セルを解除して同値を補完
        for merged in list(ws.merged_cells.ranges):
            min_row, min_col = merged.min_row, merged.min_col
            top_val = ws.cell(min_row, min_col).value
            ws.unmerge_cells(str(merged))
            for row in ws.iter_rows(
                min_row=min_row, max_row=merged.max_row,
                min_col=min_col, max_col=merged.max_col,
            ):
                for cell in row:
                    if cell.value is None:
                        cell.value = top_val

        data = [[cell.value for cell in row] for row in ws.iter_rows()]
        if not data:
            continue
        df = pd.DataFrame(data[1:], columns=data[0])
        parts.append(f"## シート: {sheet_name}\n{df.to_markdown(index=False)}")
    return "\n\n".join(parts)


def _parse_docling(path: Path) -> str:
    from docling.document_converter import DocumentConverter

    converter = DocumentConverter()
    result = converter.convert(str(path))
    return result.document.export_to_markdown()


def _parse_text(path: Path) -> str:
    for enc in ("utf-8", "cp932", "utf-16"):
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    return path.read_bytes().decode("utf-8", errors="replace")


def _parse_sync(file_path: str) -> str:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls", ".xlsm"):
        return _parse_excel(path)
    elif suffix in (".pdf", ".docx", ".pptx"):
        return _parse_docling(path)
    else:
        return _parse_text(path)


async def parse_file(file_path: str) -> str:
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, _parse_sync, file_path)
